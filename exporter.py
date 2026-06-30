import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def export_to_excel(master_df):
    """Export master dataframe to Excel with two sheets: Master_Clean and MAT_Summary."""

    output = BytesIO()

    # Define identifier columns
    id_cols = [
        "Category", "Region", "State", "Zone", "Urban_Rural",
        "TG_Segment", "Flag", "Format", "Grammage", "SU",
        "Brand_Name", "Brand_SKU_Item", "Individual_Factor"
    ]

    # Sort metric columns by type then period
    def sort_key(col):
        if "__" not in col:
            return (99, col)
        parts = col.split("__")
        metric = parts[0]
        period = parts[1] if len(parts) > 1 else ""
        order = {
            "HH": 0, "Vol": 1, "Val": 2,
            "Avg Cons": 3, "Avg FOP": 4, "Avg POC": 5, "Avg NOP": 6,
            "Units": 7, "Avg PPU": 8, "Units Estd": 9,
            "Sales Derived": 10, "Variance": 11,
            "Value MS%": 12, "Units MS%": 13
        }
        return (order.get(metric, 99), period)

    metric_cols = sorted(
        [c for c in master_df.columns if c not in id_cols],
        key=sort_key
    )

    available_id_cols = [c for c in id_cols if c in master_df.columns]
    final_cols = available_id_cols + metric_cols
    master_df = master_df[[c for c in final_cols if c in master_df.columns]]

    # MAT only columns for summary sheet
    mat_id_cols = available_id_cols
    mat_metric_cols = [c for c in metric_cols if "MAT" in c or "Mar" in c]
    mat_cols = mat_id_cols + mat_metric_cols
    mat_df = master_df[[c for c in mat_cols if c in master_df.columns]]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        master_df.to_excel(writer, sheet_name="Master_Clean", index=False)
        mat_df.to_excel(writer, sheet_name="MAT_Summary", index=False)

        workbook = writer.book
        for sheet_name in ["Master_Clean", "MAT_Summary"]:
            ws = workbook[sheet_name]
            format_sheet(ws, id_cols)

    output.seek(0)
    return output

def format_sheet(ws, id_col_names):
    """Apply formatting to worksheet."""

    # Colors
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    id_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    row_fill_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    row_fill_2 = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    headers = [cell.value for cell in ws[1]]

    # Format headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = row_fill_1 if row_idx % 2 == 0 else row_fill_2
        for cell in row:
            col_name = headers[cell.column - 1] if cell.column - 1 < len(headers) else ""
            if col_name in id_col_names:
                cell.fill = id_fill
                cell.font = Font(size=9)
            else:
                cell.fill = fill
                cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze panes
    ws.freeze_panes = "N2"

    # Column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20 if col_idx <= 13 else 15

    # Header row height
    ws.row_dimensions[1].height = 35
